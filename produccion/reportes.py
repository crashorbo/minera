from datetime import datetime

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def reporte_produccion(produccion):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename={date}-reporte-produccion.xlsx'.format(
        date=datetime.now().strftime('%Y%m%d'), )

    titulo = NamedStyle(name="titulo")
    titulo.font = Font(bold=True, size=10, color='ffffff')
    titulo.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,)
    titulo.fill = PatternFill(
        start_color='244062',
        end_color='244062',
        fill_type='solid',
    )

    titulo1 = NamedStyle(name="titulo1")
    titulo1.font = Font(bold=True, size=10, color='000000')
    titulo1.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,)
    titulo1.fill = PatternFill(
        start_color='FFFF00',
        end_color='FFFF00',
        fill_type='solid',
    )

    titulo2 = NamedStyle(name="titulo2")
    titulo2.font = Font(bold=True, size=10, color='000000')
    titulo2.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    titulo2.fill = PatternFill(
        start_color='C5D9F1',
        end_color='C5D9F1',
        fill_type='solid',
    )

    wb = Workbook()

    wb.add_named_style(titulo)

    sheet = wb.active

    sheet['C2'] = "FECHA"
    sheet['C2'].style = titulo
    sheet['D2'] = produccion.created.strftime("%d/%m/%Y")
    sheet['E2'] = "DESTINO"
    sheet['E2'].style = titulo
    sheet['F2'] = produccion.destino.nombre if produccion.destino else ""
    sheet['C3'] = "LEY MINIMA"
    sheet['C3'].style = titulo
    sheet['D3'] = produccion.minima_ley
    sheet['E3'] = "LEY MAXIMA"
    sheet['E3'].style = titulo
    sheet['F3'] = produccion.maxima_ley
    sheet['C4'] = "TOTAL TMS"
    sheet['C4'].style = titulo
    sheet['D4'] = produccion.tms
    sheet['E4'] = "TOTAL CARGAS"
    sheet['E4'].style = titulo
    sheet['F4'] = produccion.total_cargas

    sheet['A6'] = "Numero Boleta"
    sheet['A6'].style = titulo
    sheet['B6'] = "Fecha Pesaje"
    sheet['B6'].style = titulo
    sheet['C6'] = "Procedencia"
    sheet['C6'].style = titulo
    sheet['D6'] = "Tipo Carga"
    sheet['D6'].style = titulo2
    sheet['E6'] = "Peso Neto"
    sheet['E6'].style = titulo
    sheet['F6'] = "TMS"
    sheet['F6'].style = titulo2
    sheet['G6'] = "Au(g/Tn)"
    sheet['G6'].style = titulo1
    sheet['H6'] = "Estado"
    sheet['H6'].style = titulo2
    sheet['I6'] = "Color Paleta"
    sheet['I6'].style = titulo2

    column_dimensions = sheet.column_dimensions['A']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['B']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['C']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['D']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['E']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['F']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['G']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['H']
    column_dimensions.width = 20
    column_dimensions = sheet.column_dimensions['I']
    column_dimensions.width = 20

    for carga in produccion.carga_set.all():
        sheet.append((carga.numero, carga.created.strftime("%d/%m/%Y"), carga.origen.nombre, carga.tipo_carga, carga.peso_neto_tn, carga.tms_pagar, carga.au,
                      'PAGADO' if carga.pagado else 'POR PAGAR' if carga.liquido_pagable > 0 else 'NO PAGAR', '{} {}'.format(carga.numero_paleta, carga.color)))

    wb.save(response)

    return response
