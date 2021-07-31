from io import BytesIO
import re
import locale
from datetime import datetime

from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics import renderPDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.urls import reverse_lazy


class ReporteGenerador:
    def __init__(self, codigos):
        self.__codigos = codigos

    def __encabezado(self, canvas, doc):
        canvas.saveState()
        canvas.restoreState()

    def __pie(self, canvas, doc):
        canvas.saveState()
        canvas.restoreState()

    def generar_reporte_generador(self):
        hoy = datetime.now()
        response = HttpResponse(content_type='application/pdf')
        # se crea el nombre del archivo de la descarga pdf
        pdf_name = 'receta_{}.pdf'.format(hoy.strftime("%Y%m%d"))
        response['Content-Disposition'] = 'inline; filename={}'.format(
            pdf_name)
        # se crea el buffer de memoria para generar el documento
        buffer = BytesIO()
        # configuracion de documento base para la plantilla
        doc = BaseDocTemplate(buffer,
                              pagesize=letter,
                              leftMargin=1*cm,
                              rightMargin=1*cm,
                              topMargin=1*cm,
                              bottomMargin=1*cm
                              )
        frame0 = Frame(doc.leftMargin, doc.bottomMargin,
                       doc.width, doc.height, showBoundary=0, id='bordeNormal')
        doc.addPageTemplates([PageTemplate(
            id='principal', frames=frame0, onPage=self.__encabezado, onPageEnd=self.__pie), ])

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='centered',
                                  alignment=TA_CENTER, fontSize=11, fontName="Helvetica-Bold", leading=20,))

        data = []
        for codigo in self.__codigos:
            data.append((Paragraph(codigo.numero, styles['centered']), Paragraph(codigo.cod_externo, styles['centered']),
                         Paragraph(codigo.cod_proveedor, styles['centered']), Paragraph(codigo.cod_testigo, styles['centered']), Paragraph(codigo.cod_bolsa, styles['centered'])))

        tabla = Table(data, colWidths=[
                      4*cm, 4*cm, 4*cm, 4*cm, 4*cm], rowHeights=1.1*cm)
        tabla.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]
        ))
        story = []
        story.append(tabla)
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    def generar_codigos_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
        response['Content-Disposition'] = 'attachment; filename={date}-codigos.xlsx'.format(
            date=datetime.now().strftime('%Y%m%d'), )

        titulo = NamedStyle(name="titulo")
        titulo.font = Font(bold=True, size=10, color='ffffff')
        titulo.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,)
        titulo.fill = PatternFill(
            start_color='244062',
            end_color='244062',
            fill_type='solid',
        )

        wb = Workbook()

        wb.add_named_style(titulo)

        sheet = wb.active

        column_dimensions = sheet.column_dimensions['A']
        column_dimensions.width = 20
        column_dimensions = sheet.column_dimensions['B']
        column_dimensions.width = 20
        column_dimensions = sheet.column_dimensions['C']
        column_dimensions.width = 20
        column_dimensions = sheet.column_dimensions['D']
        column_dimensions.width = 20
        column_dimensions = sheet.column_dimensions['E']
        column_dimensions.width = 20

        sheet['A1'] = "CODIGO PRINCIPAL"
        sheet['A1'].style = titulo
        sheet['B1'] = "CODIGO EXTERNO"
        sheet['B1'].style = titulo
        sheet['C1'] = "CODIGO PROVEEDOR"
        sheet['C1'].style = titulo
        sheet['D1'] = "CODIGO TESTIGO"
        sheet['D1'].style = titulo
        sheet['E1'] = "CODIGO BOLSA"
        sheet['E1'].style = titulo

        for codigo in self.__codigos:
            sheet.append((codigo.numero, codigo.cod_externo,
                          codigo.cod_proveedor, codigo.cod_testigo, codigo.cod_bolsa))
        wb.save(response)

        return response
