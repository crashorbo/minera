from .models import Parametro
import datetime
from openpyxl import load_workbook
from proveedor.models import Proveedor
from conductor.models import Vehiculo, Conductor


def expedido(arg):
    switcher = {
        'ORURO': 'OR',
        'COCHABAMBA': 'CB',
        'La PAZ': 'LP'
    }

    return switcher.get(arg, '')


def activo(arg):
    switcher = {
        'ACTIVO': False,
        'INACTIVO': True
    }

    return switcher.get(arg, True)


def get_numeracion(tipo):
    today = datetime.date.today()
    try:
        parametro = Parametro.objects.get(tipo=tipo, nombre=today.year)
        parametro.valor = int(parametro.valor)+1
        parametro.save()
    except Parametro.DoesNotExist:
        parametro = Parametro(tipo=tipo, nombre=today.year, valor=1)
        parametro.save()
    return transformar_numeracion(str(parametro.nombre), parametro.valor)


def transformar_numeracion(nombre, valor):
    nombre_corto = nombre[2:4]
    aux = '{}{}'.format(nombre_corto, ('0'*6+str(valor))[-6:])
    return aux


def importar_proveedores():
    wbo = load_workbook(filename='BDATOS.xlsx')
    wso = wbo['PROVEEDORES']
    cells = wso['A2':'I510']

    for c1, c2, c3, c4, c5, c6, c7, c8, c9 in cells:
        if (c1.value):
            proveedor = Proveedor()
            proveedor.apellidos = '{} {}'.format(c2.value, c3.value)
            proveedor.nombres = '{}'.format(c5.value)
            proveedor.numero_documento = '{} {}'.format(
                c7.value, expedido(c8.value))
            proveedor.deleted = activo(c9.value)
            print('{} {} {} {}'.format(c1.value, proveedor.apellidos,
                                       proveedor.nombres, proveedor.deleted))
            proveedor.save()

    return wbo.sheetnames


def importar_vehiculos():
    wbo = load_workbook(filename='BDATOS.xlsx', data_only=True)
    wso = wbo['VEHICULOS']
    cells = wso['A2':'I60']

    for c1, c2, c3, c4, c5, c6, c7, c8, c9 in cells:
        if (c1.value):
            vehiculo = Vehiculo()
            vehiculo.placa = '{}'.format(c2.value)
            vehiculo.apellidos = '{}'.format(c9.value)
            vehiculo.nombres = '{}'.format(c7.value)
            vehiculo.modelo = '{}'.format(c3.value)
            vehiculo.descripcion = '{} {}'.format(c3.value, c4.value)
            vehiculo.save()
    return wbo.sheetnames


def importar_conductores():
    wbo = load_workbook(filename='BDATOS.xlsx', data_only=True)
    wso = wbo['CONDUCTORES']
    cells = wso['B2':'F90']

    for c1, c2, c3, c4, c5 in cells:
        if (c1.value):
            conductor = Conductor()
            conductor.apellidos = '{} {}'.format(c3.value, c4.value)
            conductor.nombres = '{}'.format(c5.value)

            conductor.save()
    return wbo.sheetnames
