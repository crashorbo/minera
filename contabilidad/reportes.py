from io import BytesIO
import re
import locale
from datetime import date, datetime
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter


from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter, portrait
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics import renderPDF

from django.http import HttpResponse
from django.urls import reverse_lazy
from reportlab.platypus.flowables import PageBreak

from pesaje.templatetags.pesaje_tags import numero_decimal
from .utils import NumeroLiteral, FechaLiteral


class ReporteContabilidad:
    def __init__(self, cargas):
        self.__cargas = cargas

    def reporte_boleta(self):
        today = datetime.now()
        response = HttpResponse(content_type='application/pdf')
        # se crea el nombre del archivo de la descarga pdf
        pdf_name = 'peso_bruto_{}.pdf'.format(today.strftime("%Y%m%d"))
        response['Content-Disposition'] = 'inline; filename={}'.format(
            pdf_name)
        # se crea el buffer de memoria para generar el documento
        buffer = BytesIO()

        pdftam = (8*cm, 12*cm)
        pdf = canvas.Canvas(buffer, pagesize=pdftam)
        for carga in self.__cargas:
            pdf.setFont('Helvetica-Bold', 10)
            pdf.drawCentredString(115, 300, "DATOS DE ENTREGA")
            pdf.setFont('Helvetica-Bold', 7)
            pdf.drawString(25, 286, "Nº Boleta:")
            pdf.drawString(25, 274, "Fecha:")
            pdf.drawString(25, 262, "Cotizacion:")
            pdf.drawCentredString(115, 234, 'Descripcion')
            pdf.drawString(25, 162, 'Valor reposicion')
            pdf.drawRightString(205, 162, '{}'.format(
                numero_decimal(carga.valor_reposicion)))
            pdf.drawString(25, 66, 'Valor descuentos')
            pdf.drawRightString(205, 66, '{}'.format(
                numero_decimal(carga.penalizacion_cu_soluble + carga.anticipo + carga.equipo_pesado + carga.balanza + carga.volqueta + \
                    carga.analisis_laboratorio + carga.otros_descuentos + carga.retencion_acuerdo))) 

            pdf.drawString(25, 43, 'Valor neto s/descuentos')
            pdf.drawRightString(205, 43, '{}'.format(
                numero_decimal(carga.liquido_pagable)))
            pdf.setFont('Helvetica', 6)
            pdf.drawString(25, 56, '{:.0f}'.format(carga.regalia))
            pdf.setFont('Helvetica', 7)
            pdf.drawRightString(205, 286, 'P{}'.format(carga.numero))
            pdf.drawRightString(205, 274, today.strftime("%d/%m/%Y"))
            pdf.drawRightString(205, 262, '{:.0f}'.format(carga.cotizacion))
            pdf.drawCentredString(115, 250, '{} {}'.format(
                carga.proveedor.apellidos, carga.proveedor.nombres))
            pdf.drawString(25, 222, 'Ton Humeda')
            pdf.drawRightString(205, 222, '{}'.format(
                numero_decimal(carga.peso_neto_tn)))
            pdf.drawString(25, 210, 'Humedad %')
            pdf.drawRightString(205, 210, '{:.0f}'.format(carga.h2o))
            pdf.drawString(25, 198, 'Ton Seco')
            pdf.drawRightString(205, 198, '{}'.format(
                numero_decimal(carga.tms_pagar)))
            pdf.drawString(25, 186, 'Ley')
            pdf.drawRightString(205, 186, '{}'.format(
                numero_decimal(carga.au)))
            pdf.drawString(25, 174, 'Finos recuper. Gr. - 60')
            pdf.drawRightString(205, 174, '{}'.format(
                numero_decimal(carga.finos_gr_recup)))
            pdf.drawString(25, 150, 'Cobre soluble')
            pdf.drawRightString(205, 150, '{}'.format(
                numero_decimal(carga.penalizacion_cu_soluble)))
            pdf.drawString(25, 138, 'Anticipos')
            pdf.drawRightString(205, 138, '{}'.format(
                numero_decimal(carga.anticipo)))
            pdf.drawString(25, 126, 'Pala (carguio)')
            pdf.drawRightString(205, 126, '{}'.format(
                numero_decimal(carga.equipo_pesado)))
            pdf.drawString(25, 114, 'Balanza (peso)')
            pdf.drawRightString(205, 114, '{}'.format(
                numero_decimal(carga.balanza)))
            pdf.drawString(25, 102, 'Volqueta (Acarreo)')
            pdf.drawRightString(205, 102, '{}'.format(
                numero_decimal(carga.volqueta)))
            pdf.drawString(25, 90, 'Laboratorio (Analisis)')
            pdf.drawRightString(205, 90, '{}'.format(
                numero_decimal(carga.analisis_laboratorio)))
            pdf.drawString(25, 78, 'Descuentos p/terceros')
            pdf.drawRightString(205, 78, '{}'.format(
                numero_decimal(carga.otros_descuentos)))

            pdf.line(20, 243, 210, 243)
            pdf.line(20, 231, 210, 231)
            pdf.line(20, 243, 20, 231)
            pdf.line(210, 243, 210, 231)

            pdf.line(20, 52, 210, 52)
            pdf.line(20, 40, 210, 40)
            pdf.line(20, 52, 20, 40)
            pdf.line(210, 52, 210, 40)

            pdf.setDash(0.5, 1.5)

            pdf.line(20, 295, 210, 295)
            pdf.line(20, 283, 210, 283)
            pdf.line(20, 271, 210, 271)
            pdf.line(20, 259, 210, 259)
            pdf.line(20, 247, 210, 247)
            pdf.line(20, 247, 20, 295)
            pdf.line(210, 247, 210, 295)

            pdf.line(20, 219, 210, 219)
            pdf.line(20, 207, 210, 207)
            pdf.line(20, 195, 210, 195)
            pdf.line(20, 183, 210, 183)
            pdf.line(20, 171, 210, 171)
            pdf.line(20, 159, 210, 159)
            pdf.line(20, 147, 210, 147)
            pdf.line(20, 135, 210, 135)
            pdf.line(20, 123, 210, 123)
            pdf.line(20, 111, 210, 111)
            pdf.line(20, 99, 210, 99)
            pdf.line(20, 87, 210, 87)
            pdf.line(20, 75, 210, 75)
            pdf.line(20, 63, 210, 63)
            pdf.line(20, 231, 20, 63)
            pdf.line(210, 231, 210, 63)
            pdf.showPage()

        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response


class ReporteComprobante():
    def __init__(self, cargas):
        self.__cargas = cargas

    def __encabezado(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 10)
        canvas.drawCentredString(6*cm, 13.5*cm, 'EMPRESA MINERA COMUNITARIA')
        canvas.drawCentredString(6*cm, 13*cm, '"INCA SAYAÑA" S.A.')
        canvas.setFont('Helvetica-Bold', 16)
        canvas.drawCentredString(
            12*cm, 12*cm, 'COMPROBANTE DE CAJA - EGRESOS')
        canvas.line(6.8*cm, 11.9*cm, 17.2*cm, 11.9*cm)
        canvas.restoreState()

    def __pie(self, canvas, doc):
        canvas.saveState()
        canvas.restoreState()

    def generar_comprobantes(self):
        pass

    def generar_comprobante(self):
        hoy = datetime.now()
        response = HttpResponse(content_type='application/pdf')
        # se crea el nombre del archivo de la descarga pdf
        pdf_name = 'comprobante_{}.pdf'.format(hoy.strftime("%Y%m%d"))
        response['Content-Disposition'] = 'inline; filename={}'.format(
            pdf_name)
        # se crea el buffer de memoria para generar el documento
        buffer = BytesIO()

        pdftam = (22*cm, 14.5*cm)
        # configuracion de documento base para la plantilla
        doc = BaseDocTemplate(buffer,
                              pagesize=pdftam,
                              leftMargin=1*cm,
                              rightMargin=1*cm,
                              topMargin=1*cm,
                              bottomMargin=1*cm
                              )
        frame0 = Frame(doc.leftMargin + 2*cm, doc.bottomMargin,
                       doc.width-2*cm, doc.height - 2.5*cm, showBoundary=1, id='bordeNormal')
        doc.addPageTemplates([PageTemplate(
            id='principal', frames=frame0, onPage=self.__encabezado, onPageEnd=self.__pie), ])

        story = []

        for carga in self.__cargas:
            numero_literal = NumeroLiteral()

            data0 = [['FECHA', '', '', 'Nro:', 'P{}'.format(carga.numero)],
                     ['{}'.format(hoy.day), '{}'.format(hoy.month), '{}'.format(hoy.year), '', '']]

            t0 = Table(data0, [1.5*cm, 1.5*cm, 1.5*cm,
                               10*cm, 2.9*cm], [0.8*cm, 0.8*cm])

            t0.setStyle(TableStyle([
                ('GRID', (0, 0), (2, 1), 0.5, colors.black),
                ('SPAN', (0, 0), (2, 0)),
                ('FONT', (0, 0), (2, 0), 'Helvetica-Bold', 10),
                ('VALIGN', (0, 0), (4, 1), 'MIDDLE'),
                ('ALIGN', (0, 0), (2, 1), 'CENTER'),
                ('ALIGN', (3, 0), (3, 1), 'RIGHT'),
                ('FONT', (3, 0), (3, 1), 'Helvetica-Bold', 14),
                ('FONT', (4, 0), (4, 1), 'Helvetica', 14),
                ('SPAN', (3, 0), (3, 1)),
                ('SPAN', (4, 0), (4, 1)),
            ]))

            data1 = [['Pagado a:', '{} {}'.format(carga.proveedor.apellidos, carga.proveedor.nombres)],
                     ['La Suma de:', Paragraph('{} 00/100 BOLIVIANOS.'.format(
                         numero_literal.main(int(carga.liquido_pagable))))],
                     ['Por Concepto:',
                      'CARGA MINERALIZADA: P{} - {} - COT:{}'.format(carga.numero, carga.created.strftime("%d/%m/%Y"), int(carga.cotizacion))],
                     ['', 'TMH:{} - HUM%:{} - TMS:{} - LEY:{} -  Fi.Rec.:{} - V.REP:{}'.format(
                         carga.peso_neto_tn, int(carga.h2o), carga.tms_pagar, carga.au, carga.finos_gr_recup, int(carga.valor_reposicion))],
                     ['', 'PP:{} - ANT:{} - EQP:{} - BAL:{} - VOL:{} - LAB:{} DSC:{} - V.DSC:{}'.format(
                         int(carga.penalizacion_cu_soluble), int(
                             carga.anticipo), int(carga.equipo_pesado),
                         int(carga.balanza), int(carga.volqueta), int(carga.analisis_laboratorio), int(carga.otros_descuentos), int(carga.penalizacion_cu_soluble + carga.anticipo + carga.equipo_pesado + carga.balanza + carga.volqueta + \
                            carga.analisis_laboratorio + carga.otros_descuentos + carga.retencion_acuerdo))],
                     ]
            t1 = Table(data1, [3.2*cm, 14.2*cm],
                       [0.6*cm, 1.2*cm, 0.6*cm, 0.6*cm, 0.6*cm])

            t1.setStyle(TableStyle([
                ('VALIGN', (0, 0), (1, 0), 'MIDDLE'),
                ('VALIGN', (0, 1), (4, 2), 'TOP'),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                ('FONT', (0, 0), (0, 4), 'Helvetica-Bold', 10),
            ]))

            data2 = [['Bs:', '{}'.format(numero_decimal(carga.liquido_pagable)), ''],
                     ['Efectivo:', '', ''],
                     ['Cheque No.:', '', ''],
                     ['Banco:', '', ''],
                     ['Cuenta Contable:', '', 'Firma Beneficiario'],
                     ['', '', 'C.I./NIT:'],
                     ]

            t2 = Table(data2, [3.2*cm, 5.5*cm, 8.7*cm], 6*[0.6*cm])

            t2.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                ('LINEBEFORE', (2, 0), (2, 6), 1, colors.black),
                ('ALIGN', (2, 4), (2, 5), 'CENTER'),
                ('RIGHTPADDING', (2, 5), (2, 5), 100),
                ('FONT', (0, 0), (0, 4), 'Helvetica-Bold', 10),
            ]))

            story.append(t0)
            story.append(Spacer(0, 5))
            story.append(t1)
            story.append(Spacer(0, 5))
            story.append(t2)
            story.append(PageBreak())

        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response


class ReporteExcel():

    def __init__(self, carga):
        self.__carga = carga

    def reporte_por_pagar(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
        response['Content-Disposition'] = 'attachment; filename={date}-porpagar.xlsx'.format(
            date=datetime.now().strftime('%Y%m%d'), )

        titulo = NamedStyle(name="titulo")
        titulo.font = Font(bold=True, size=14, color='ffffff')
        titulo.alignment = Alignment(
            horizontal="center", vertical="center")
        titulo.fill = PatternFill(
            start_color='0070C0',
            end_color='0070C0',
            fill_type='solid',
        )

        monto = NamedStyle(name="monto")
        monto.font = Font(bold=True, size=14,)
        monto.alignment = Alignment(horizontal="right", vertical="center")
        wb = Workbook()

        wb.add_named_style(titulo)

        sheet = wb.active

        column_dimensions = sheet.column_dimensions['C']
        column_dimensions.width = 60

        sheet['A1'] = "Numero"
        sheet['A1'].style = titulo
        sheet['B1'] = "Fecha"
        sheet['B1'].style = titulo
        sheet['C1'] = "Proveedor"
        sheet['C1'].style = titulo
        sheet['D1'] = "Monto"
        sheet['D1'].style = titulo

        liquido_pagable = 0

        for row in self.__carga:
            if row.liquido_pagable > 0:
                liquido_pagable = liquido_pagable + row.liquido_pagable
                sheet.append((row.numero, row.created.strftime("%d/%m/%Y"), '{} {}'.format(
                    row.proveedor.apellidos, row.proveedor.nombres), row.liquido_pagable))

        max_row = sheet.max_row

        total_liquido_pagable_descr_cell = sheet.cell(
            row=max_row + 2, column=sheet.max_column - 1)
        total_liquido_pagable_descr_cell.value = "Total"

        total_liquido_pagable = sheet.cell(
            row=max_row + 2, column=sheet.max_column)
        total_liquido_pagable.value = liquido_pagable

        total_liquido_pagable_descr_cell.style = titulo
        total_liquido_pagable.style = monto

        for i in range(1, sheet.max_column+1):
            sheet.column_dimensions[get_column_letter(i)].bestFit = True
            sheet.column_dimensions[get_column_letter(i)].auto_size = True

        wb.save(response)

        return response


class ReporteCargasPagadas():

    def __init__(self, cargas):
        self.__cargas = cargas

    def reporte_cargas_pagadas(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
        response['Content-Disposition'] = 'attachment; filename={date}-cargas-pagadas.xlsx'.format(
            date=datetime.now().strftime('%Y%m%d'), )

        titulo = NamedStyle(name="titulo")
        titulo.font = Font(bold=True, size=10, color='ffffff')
        titulo.alignment = Alignment(
            horizontal="center", vertical="center")
        titulo.fill = PatternFill(
            start_color='0070C0',
            end_color='0070C0',
            fill_type='solid',
        )

        wb = Workbook()

        wb.add_named_style(titulo)

        sheet = wb.active

        column_dimensions = sheet.column_dimensions['C']
        column_dimensions.width = 60

        column_dimensions = sheet.column_dimensions['D']
        column_dimensions.width = 16

        column_dimensions = sheet.column_dimensions['E']
        column_dimensions.width = 16

        column_dimensions = sheet.column_dimensions['F']
        column_dimensions.width = 16

        column_dimensions = sheet.column_dimensions['G']
        column_dimensions.width = 40

        sheet['A1'] = "NUM. BOL."
        sheet['A1'].style = titulo
        sheet['B1'] = "FECHA"
        sheet['B1'].style = titulo
        sheet['C1'] = "CONCEPTO"
        sheet['C1'].style = titulo
        sheet['D1'] = "EGRESO (Bs)"
        sheet['D1'].style = titulo
        sheet['E1'] = "EGRESO NETO (Bs)"
        sheet['E1'].style = titulo
        sheet['F1'] = "CODIGO CONTABLE"
        sheet['F1'].style = titulo
        sheet['G1'] = "NOMBRE CUENTA"
        sheet['G1'].style = titulo

        for carga in self.__cargas:
            sheet.append((carga.numero, carga.fecha_pago.strftime("%d/%m/%Y"), 'PESAJE - {} {}; {} {}'.format(carga.numero, FechaLiteral().fecha_literal(carga.fecha_pago),
                                                                                                              carga.proveedor.apellidos, carga.proveedor.nombres), carga.liquido_pagable, carga.liquido_pagable, '2110301', 'PROVEEDORES CARGA MINERALIZADA'))
        wb.save(response)

        return response


class ReporteCargasGeneral():
    def __init__(self, cargas):
        self.__cargas = cargas

    def reporte_cargas_general(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
        response['Content-Disposition'] = 'attachment; filename={date}-cargas-pagadas.xlsx'.format(
            date=datetime.now().strftime('%Y%m%d'), )

        titulo = NamedStyle(name="titulo")
        titulo.font = Font(bold=True, size=10, color='ffffff')
        titulo.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,)
        titulo.fill = PatternFill(
            start_color='244062',
            end_color='244062',
            fill_type='solid',
        )

        titulo1 = NamedStyle(name="titulo1")
        titulo1.font = Font(bold=True, size=10, color='000000')
        titulo1.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,)
        titulo1.fill = PatternFill(
            start_color='FFFF00',
            end_color='FFFF00',
            fill_type='solid',
        )

        titulo2 = NamedStyle(name="titulo2")
        titulo2.font = Font(bold=True, size=10, color='000000')
        titulo2.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo2.fill = PatternFill(
            start_color='C5D9F1',
            end_color='C5D9F1',
            fill_type='solid',
        )

        titulo3 = NamedStyle(name="titulo3")
        titulo3.font = Font(bold=True, size=10, color='000000')
        titulo3.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo3.fill = PatternFill(
            start_color='FCD5B4',
            end_color='FCD5B4',
            fill_type='solid',
        )

        titulo4 = NamedStyle(name="titulo4")
        titulo4.font = Font(bold=True, size=10, color='000000')
        titulo4.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo4.fill = PatternFill(
            start_color='D8E4BC',
            end_color='D8E4BC',
            fill_type='solid',
        )

        titulo5 = NamedStyle(name="titulo5")
        titulo5.font = Font(bold=True, size=10, color='000000')
        titulo5.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo5.fill = PatternFill(
            start_color='C4D79B',
            end_color='C4D79B',
            fill_type='solid',
        )

        titulo6 = NamedStyle(name="titulo6")
        titulo6.font = Font(bold=True, size=10, color='000000')
        titulo6.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo6.fill = PatternFill(
            start_color='FABF8F',
            end_color='FABF8F',
            fill_type='solid',
        )

        titulo7 = NamedStyle(name="titulo7")
        titulo7.font = Font(bold=True, size=10, color='000000')
        titulo7.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        titulo7.fill = PatternFill(
            start_color='00FFFF',
            end_color='00FFFF',
            fill_type='solid',
        )

        wb = Workbook()

        wb.add_named_style(titulo)

        sheet = wb.active

        column_dimensions = sheet.column_dimensions['D']
        column_dimensions.width = 40
        column_dimensions = sheet.column_dimensions['E']
        column_dimensions.width = 40
        column_dimensions = sheet.column_dimensions['AQ']
        column_dimensions.width = 40
        column_dimensions = sheet.column_dimensions['AU']
        column_dimensions.width = 20
        column_dimensions = sheet.column_dimensions['AV']
        column_dimensions.width = 40

        sheet['A1'] = "NUM. BOL."
        sheet['A1'].style = titulo
        sheet['B1'] = "FECHA"
        sheet['B1'].style = titulo
        sheet['C1'] = "PLACA VEHICULO"
        sheet['C1'].style = titulo
        sheet['D1'] = "NOMBRE TRANSPORTISTA"
        sheet['D1'].style = titulo
        sheet['E1'] = "NOMBRE PROPIETARIO"
        sheet['E1'].style = titulo
        sheet['F1'] = "PROCEDENCIA"
        sheet['F1'].style = titulo
        sheet['G1'] = "DESTINO CARGA"
        sheet['G1'].style = titulo
        sheet['H1'] = "PESO BRUTO"
        sheet['H1'].style = titulo
        sheet['I1'] = "PESO TARA"
        sheet['I1'].style = titulo
        sheet['J1'] = "PESO NETO"
        sheet['J1'].style = titulo
        sheet['K1'] = "HORA PESAJE"
        sheet['K1'].style = titulo
        sheet['L1'] = "TIPO CARGA"
        sheet['L1'].style = titulo1
        sheet['M1'] = "Au (g/Tn)"
        sheet['M1'].style = titulo1
        sheet['N1'] = "H2O (%)"
        sheet['N1'].style = titulo1
        sheet['O1'] = "% SOBRE TAMANOS"
        sheet['O1'].style = titulo1
        sheet['P1'] = "COBRE SOLUBLE CN"
        sheet['P1'].style = titulo1
        sheet['Q1'] = "T.M. SECA"
        sheet['Q1'].style = titulo2
        sheet['R1'] = "TMS NETAS"
        sheet['R1'].style = titulo2
        sheet['S1'] = "TMS A PENALIZAR"
        sheet['S1'].style = titulo2
        sheet['T1'] = "TMS A PAGAR"
        sheet['T1'].style = titulo2
        sheet['U1'] = "FINOS GR/AU"
        sheet['U1'].style = titulo2
        sheet['V1'] = "FINOS OZ/AU"
        sheet['V1'].style = titulo2
        sheet['W1'] = "COTIZACION QUINCENA"
        sheet['W1'].style = titulo2
        sheet['X1'] = "VALOR P/CALCULO REGALIA"
        sheet['X1'].style = titulo2
        sheet['Y1'] = "RECUPER. PLANTA 60A62/GR"
        sheet['Y1'].style = titulo3
        sheet['Z1'] = "VALOR VENTA AL 94%"
        sheet['Z1'].style = titulo3
        sheet['AA1'] = "COSTO TRATAT. PLANTA 50%"
        sheet['AA1'].style = titulo3
        sheet['AB1'] = "TOTAL LIQUIDAC. PROV. Bs"
        sheet['AB1'].style = titulo3
        sheet['AC1'] = "REGALIA 4,2"
        sheet['AC1'].style = titulo4
        sheet['AD1'] = "PENALIZA Cu SOLUBLE"
        sheet['AD1'].style = titulo4
        sheet['AE1'] = "ANTICIPO"
        sheet['AE1'].style = titulo4
        sheet['AF1'] = "EQUIPO PESADO"
        sheet['AF1'].style = titulo4
        sheet['AG1'] = "BALANZA"
        sheet['AG1'].style = titulo4
        sheet['AH1'] = "VOLQUETA"
        sheet['AH1'].style = titulo4
        sheet['AI1'] = "ANALISIS LABORATORIO"
        sheet['AI1'].style = titulo4
        sheet['AJ1'] = "OTROS DESCUENTOS"
        sheet['AJ1'].style = titulo4
        sheet['AK1'] = "RETENCION 5% ACUERDO VOL."
        sheet['AK1'].style = titulo4
        sheet['AL1'] = "TOTAL DESCUENTOS"
        sheet['AL1'].style = titulo5
        sheet['AM1'] = "LIQUIDO PAGABLE"
        sheet['AM1'].style = titulo6
        sheet['AN1'] = "FECHA DE PAGO"
        sheet['AN1'].style = titulo7
        sheet['AO1'] = "NUMERO Y COLOR PALETA"
        sheet['AO1'].style = titulo7
        sheet['AP1'] = "ESTADO"
        sheet['AP1'].style = titulo7
        sheet['AQ1'] = "EQUIPO CARGUIO"
        sheet['AQ1'].style = titulo
        sheet['AR1'] = "ORO SOLUBLE"
        sheet['AR1'].style = titulo3
        sheet['AS1'] = "RATIO"
        sheet['AS1'].style = titulo3
        sheet['AT1'] = "FECHA MUESTREO"
        sheet['AT1'].style = titulo2
        sheet['AU1'] = "FECHA LABORATORIO"
        sheet['AU1'].style = titulo2
        sheet['AV1'] = "HISTORIAL LABORATORIO MODIFICACIONES"
        sheet['AV1'].style = titulo2

        for carga in self.__cargas:
            sheet.append((carga.numero, carga.created.strftime("%d/%m/%Y"), carga.vehiculo.placa, '{} {}'.format(
                carga.conductor_vehiculo.apellidos, carga.conductor_vehiculo.nombres), '{} {}'.format(carga.proveedor.apellidos, carga.proveedor.nombres), carga.origen.nombre, carga.destino.nombre,
                carga.peso_bruto, carga.peso_tara, carga.peso_neto_tn, carga.created.strftime(
                    "%H:%M:%S"), carga.tipo_carga, carga.au, carga.h2o, carga.porcentaje_tamanos, carga.cobre_soluble,
                carga.tms, carga.tms_neta, carga.tms_penalizar, carga.tms_pagar, carga.finos_gr, carga.finos_oz, carga.cotizacion, carga.calculo_regalia, carga.recu_planta, carga.valor_venta,
                carga.costo_tratamiento, carga.total_liquidacion_prov, carga.regalia, carga.penalizacion_cu_soluble, carga.anticipo, carga.equipo_pesado, carga.balanza, carga.volqueta,
                carga.analisis_laboratorio, carga.otros_descuentos, carga.retencion_acuerdo, carga.total_descuento, carga.liquido_pagable, carga.fecha_pago.strftime(
                    "%d/%m/%Y") if carga.fecha_pago else '', '{} {}'.format(carga.numero_paleta, carga.color), 'PAGADO' if carga.pagado else 'POR PAGAR' if carga.liquido_pagable > 0 else 'NO PAGAR',
                '{} {}'.format(carga.equipo_carguio.apellidos, carga.equipo_carguio.nombres) if carga.equipo_carguio else 'MANUAL', carga.oro_soluble, carga.ratio, carga.fecha_muestreo.strftime("%d/%m/%Y") if carga.fecha_muestreo else '',
                carga.fecha_laboratorio.strftime("%d/%m/%Y %H:%M") if carga.fecha_laboratorio else '', carga.fecha_laboratorio_modificaciones))

        wb.save(response)

        return response
